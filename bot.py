import logging
import json
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, filters, ContextTypes
)
from anthropic import Anthropic
import openpyxl
import os
import gspread
from google.oauth2.service_account import Credentials
from config_1 import TELEGRAM_TOKENN, CLAUDE_API_KEYY, SYSTEM_PROMPT

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# СОСТОЯНИЯ
# ─────────────────────────────────────────────
RULE1, RULE2, RULE3, RULE4, RULE5 = range(5)
NAME, PHONE, CITIZENSHIP, IIN, CONSENT = range(5, 10)

EXCEL_FILE = "students.xlsx"
client = Anthropic(api_key=CLAUDE_API_KEYY)


# ─────────────────────────────────────────────
# GOOGLE SHEETS — инициализация
# ─────────────────────────────────────────────
# Переменные окружения (добавить в .env и Railway):
#   GOOGLE_SHEET_ID  — ID таблицы из URL (между /d/ и /edit)
#   GOOGLE_CREDENTIALS_JSON — содержимое файла credentials.json (весь JSON одной строкой)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

def get_google_sheet():
    """Возвращает объект листа Google Sheets или None при ошибке."""
    try:
        creds_source = (
            os.getenv("4BOT_CREDS_JSON")
            or os.getenv("GOOGLE_CREDENTIALS_JSON")
            or os.getenv("GOOGLE_CREDS_JSON")
        )
        sheet_id = os.getenv("4BOT_SHEET_ID") or os.getenv("GOOGLE_SHEET_ID")
        if not creds_source or not sheet_id:
            logger.warning("4BOT_CREDS_JSON или 4BOT_SHEET_ID не заданы — Google Sheets отключён")
            return None

        creds_source = creds_source.strip()
        if creds_source.startswith("{"):
            creds_data = json.loads(creds_source)
        else:
            with open(creds_source, "r", encoding="utf-8") as credentials_file:
                creds_data = json.load(credentials_file)

        creds = Credentials.from_service_account_info(creds_data, scopes=SCOPES)
        gc = gspread.authorize(creds)
        spreadsheet = gc.open_by_key(sheet_id)
        # Берём первый лист; если его нет — создаём
        try:
            worksheet = spreadsheet.worksheet("Студенты")
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title="Студенты", rows=1000, cols=6)
            worksheet.append_row(
                ["Имя и фамилия", "Телефон", "ИИН / Паспорт",
                 "Telegram ID", "Username", "Дата регистрации"],
                value_input_option="RAW"
            )
        return worksheet
    except Exception as e:
        logger.error(f"Ошибка подключения к Google Sheets: {e}")
        return None


def ensure_sheet_header(worksheet):
    """Добавляет заголовок, если лист пустой."""
    if worksheet.row_count == 0 or not worksheet.row_values(1):
        worksheet.append_row(
            ["Имя и фамилия", "Телефон", "ИИН / Паспорт",
             "Telegram ID", "Username", "Дата регистрации"],
            value_input_option="RAW"
        )


# ─────────────────────────────────────────────
# СОХРАНЕНИЕ ДАННЫХ — Google Sheets + Excel (резерв)
# ─────────────────────────────────────────────

def save_student_data(data: dict):
    """Сохраняет данные студента в Google Sheets и в Excel (резервная копия)."""
    row = [
        data["name"],
        data["phone"],
        data["iin"],
        data["telegram_id"],
        data.get("username", "—"),
        data["date"]
    ]

    # — Google Sheets —
    try:
        worksheet = get_google_sheet()
        if worksheet is not None:
            ensure_sheet_header(worksheet)
            worksheet.append_row(row, value_input_option="USER_ENTERED")
            logger.info(f"Студент {data['name']} записан в Google Sheets")
    except Exception as e:
        logger.error(f"Ошибка записи в Google Sheets: {e}")

    # — Excel (резервная копия на сервере) —
    try:
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Студенты"
            ws.append(["Имя и фамилия", "Телефон", "ИИН / Паспорт",
                       "Telegram ID", "Username", "Дата регистрации"])
        ws.append(row)
        wb.save(EXCEL_FILE)
    except Exception as e:
        logger.error(f"Ошибка записи в Excel: {e}")


def is_registered(telegram_id: int) -> dict | None:
    """Проверяет регистрацию студента — сначала в Google Sheets, потом в Excel."""
    # — Проверка в Google Sheets —
    try:
        worksheet = get_google_sheet()
        if worksheet is not None:
            records = worksheet.get_all_values()
            for row in records[1:]:  # пропускаем заголовок
                if len(row) >= 4 and str(row[3]) == str(telegram_id):
                    return {"name": row[0]}
    except Exception as e:
        logger.error(f"Ошибка чтения из Google Sheets: {e}")

    # — Резерв: проверка в Excel —
    try:
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3] == telegram_id:
                    return {"name": row[0]}
    except Exception as e:
        logger.error(f"Ошибка чтения из Excel: {e}")

    return None


# ─────────────────────────────────────────────
# ТЕКСТЫ ПРАВИЛ (дословно, MarkdownV2)
# ─────────────────────────────────────────────

RULES = {
    RULE1: (
        "Привет\\!\n\n"
        "Прежде чем получить доступ к ИИ\\-агенту PRIMA TA, ознакомьтесь с "
        "условиями использования\\. Это займёт 2–3 минуты\\.\n\n"
        "*Что такое ИИ\\-агент PRIMA TA*\n\n"
        "Это образовательный инструмент для поддержки вашего обучения "
        "в рамках программы института\\.\n\n"
        "ИИ\\-агент:\n"
        "— не является терапевтом, супервизором или консультантом\n"
        "— не обладает клиническим мышлением\n"
        "— не принимает профессиональных решений\n"
        "— может содержать неточности в ответах\n\n"
        "Вы сохраняете полную ответственность за свои профессиональные решения\\.\n\n"
        "_ИИ\\-ассистент работает только в рамках материалов вашего модуля — "
        "он не заменяет преподавателя, живое обучение и супервизию\\._"
    ),
    RULE2: (
        "*Что нельзя вводить в чат с ИИ*\n\n"
        "Запрещено вводить:\n"
        "— персональные данные клиентов\n"
        "— описания реальных кейсов, по которым можно идентифицировать человека\n"
        "— записи сессий и материалы супервизий\n"
        "— любые данные третьих лиц\n\n"
        "ИИ нельзя использовать для терапии или супервизии\\.\n\n"
        "_Это этическое требование, а не формальность\\._"
    ),
    RULE3: (
        "*Конфиденциальность*\n\n"
        "Всё, что вы получаете через ИИ\\-агента — материалы, структуры, "
        "комментарии — является интеллектуальной собственностью PRIMA TA\\.\n\n"
        "Вы обязуетесь:\n"
        "— не передавать ссылку доступа третьим лицам\n"
        "— не распространять материалы и ответы ИИ\n"
        "— не воспроизводить архитектуру инструмента\n"
        "— не создавать производные продукты на его основе\n\n"
        "Обязательства действуют бессрочно\\."
    ),
    RULE4: (
        "*О рисках использования ИИ*\n\n"
        "Пожалуйста, осознайте следующее:\n\n"
        "— ИИ может ошибаться\n"
        "— ИИ может создавать иллюзию понимания там, где его нет\n"
        "— есть риск переоценить достоверность ответов\n"
        "— есть риск избыточной зависимости от инструмента\n\n"
        "Критическое мышление остаётся за вами\\."
    ),
    RULE5: (
        "*Ответственность и последствия*\n\n"
        "Вы используете ИИ\\-ассистент на собственную ответственность\\.\n\n"
        "PRIMA TA не несёт ответственности за решения, "
        "принятые на основе ответов ИИ\\.\n\n"
        "При нарушении условий соглашения PRIMA TA вправе:\n"
        "— ограничить или прекратить ваш доступ\n"
        "— исключить вас из программы\n"
        "— потребовать удаления полученных материалов\n"
        "— применить меры в соответствии с законодательством"
    ),
}


# ─────────────────────────────────────────────
# ФАЗА 0 — Правила пользования (5 шагов)
# ─────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("onboarded"):
        registered = is_registered(update.effective_user.id)
        if registered:
            context.user_data["onboarded"] = True
            context.user_data["student_name"] = registered["name"]
            context.user_data.setdefault("history", [])

    if context.user_data.get("onboarded"):
        name = context.user_data["student_name"]
        await update.message.reply_text(
            f"С возвращением, {name}! 👋\nЗадавайте вопросы по курсу."
        )
        return ConversationHandler.END

    context.user_data.clear()
    await update.message.reply_text(
        RULES[RULE1],
        parse_mode="MarkdownV2",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("Продолжить →", callback_data="to_rule2")
        ]])
    )
    return RULE1


async def rule1_to_rule2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text(
        RULES[RULE2],
        parse_mode="MarkdownV2",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("Принимаю →", callback_data="to_rule3")
        ]])
    )
    return RULE2


async def rule2_to_rule3(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text(
        RULES[RULE3],
        parse_mode="MarkdownV2",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("Принимаю →", callback_data="to_rule4")
        ]])
    )
    return RULE3


async def rule3_to_rule4(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text(
        RULES[RULE4],
        parse_mode="MarkdownV2",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("Осознаю →", callback_data="to_rule5")
        ]])
    )
    return RULE4


async def rule4_to_rule5(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text(
        RULES[RULE5],
        parse_mode="MarkdownV2",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("Принимаю →", callback_data="to_registration")
        ]])
    )
    return RULE5


async def rule5_to_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.message.reply_text(
        "Отлично\\! Теперь нам нужно зафиксировать ваши данные\\.\n\n"
        "Пожалуйста, введите ваше *имя и фамилию*:",
        parse_mode="MarkdownV2"
    )
    return NAME


# ─────────────────────────────────────────────
# ФАЗА 1 — Сбор данных
# ─────────────────────────────────────────────

async def handle_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if len(name) < 2:
        await update.message.reply_text("Пожалуйста, введите имя и фамилию.")
        return NAME
    context.user_data["student_name"] = name
    await update.message.reply_text("Введите ваш *номер телефона*:", parse_mode="Markdown")
    return PHONE


async def handle_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text.strip()
    context.user_data["student_phone"] = phone
    keyboard = [[
        InlineKeyboardButton("🇰🇿 Гражданин(ка) Казахстана", callback_data="country_kz"),
        InlineKeyboardButton("🌍 Другая страна", callback_data="country_other"),
    ]]
    await update.message.reply_text(
        "Укажите ваше гражданство:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return CITIZENSHIP


async def handle_citizenship(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "country_kz":
        context.user_data["citizenship"] = "KZ"
        await query.edit_message_text("Введите ваш *ИИН* (12 цифр):", parse_mode="Markdown")
    else:
        context.user_data["citizenship"] = "OTHER"
        await query.edit_message_text("Введите *номер вашего паспорта*:", parse_mode="Markdown")
    return IIN


async def handle_iin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    iin = update.message.text.strip()
    context.user_data["student_iin"] = iin
    await _show_consent(update.message)
    return CONSENT


async def _show_consent(message_or_query):
    privacy_text = (
        "📋 *Согласие на обработку персональных данных*\n\n"
        "Ваши данные (ФИО, телефон, ИИН/паспорт) сохраняются в базе института PRIMA TA "
        "исключительно для учёта участников курса.\n\n"
        "Данные не передаются третьим лицам и используются только в образовательных целях.\n\n"
        "Вы согласны на обработку персональных данных?"
    )
    keyboard = [[
        InlineKeyboardButton("✅ Согласен(а)", callback_data="consent_yes"),
        InlineKeyboardButton("❌ Не согласен(а)", callback_data="consent_no"),
    ]]
    markup = InlineKeyboardMarkup(keyboard)
    if hasattr(message_or_query, "reply_text"):
        await message_or_query.reply_text(privacy_text, parse_mode="Markdown", reply_markup=markup)
    else:
        await message_or_query.edit_message_text(privacy_text, parse_mode="Markdown", reply_markup=markup)


async def handle_consent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "consent_no":
        keyboard = [[InlineKeyboardButton("↩️ Вернуться к согласию", callback_data="consent_return")]]
        await query.edit_message_text(
            "⚠️ Без согласия на обработку персональных данных доступ к боту невозможен.\n\n"
            "Если хотите продолжить — вернитесь и дайте согласие.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return CONSENT

    if query.data == "consent_return":
        await _show_consent(query)
        return CONSENT

    # Сохраняем данные
    user = query.from_user
    save_student_data({
        "name": context.user_data["student_name"],
        "phone": context.user_data["student_phone"],
        "iin": context.user_data["student_iin"],
        "telegram_id": user.id,
        "username": f"@{user.username}" if user.username else "—",
        "date": datetime.now().strftime("%d.%m.%Y %H:%M")
    })

    context.user_data["history"] = []
    context.user_data["onboarded"] = True

    name = context.user_data["student_name"]
    await query.edit_message_text(
        f"Отлично, {name}! Данные сохранены ✅\n\n"
        "Теперь вы можете задавать вопросы по курсу.\n"
        "Я здесь, чтобы помочь вам разобраться в темах, порефлексировать "
        "и углубить понимание.\n\n"
        "С чего начнём?"
    )
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Диалог прерван. Вернитесь в любой момент — напишите /start"
    )
    return ConversationHandler.END


# ─────────────────────────────────────────────
# ФАЗА 2 — Диалог с Claude
# ─────────────────────────────────────────────

async def chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("onboarded"):
        await update.message.reply_text(
            "Пожалуйста, сначала пройдите регистрацию — напишите /start"
        )
        return

    user_message = update.message.text
    history: list = context.user_data.setdefault("history", [])
    history.append({"role": "user", "content": user_message})

    await context.bot.send_chat_action(
        chat_id=update.effective_chat.id, action="typing"
    )

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1000,
        system=SYSTEM_PROMPT,
        messages=history
    )
    assistant_reply = response.content[0].text
    history.append({"role": "assistant", "content": assistant_reply})
    await update.message.reply_text(assistant_reply)


# ─────────────────────────────────────────────
# КОМАНДА /reset
# ─────────────────────────────────────────────

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("onboarded"):
        context.user_data["history"] = []
        await update.message.reply_text(
            "История диалога очищена 🔄\nНачинаем разговор заново."
        )
    else:
        await update.message.reply_text("Напишите /start чтобы начать.")


# ─────────────────────────────────────────────
# ЗАПУСК
# ─────────────────────────────────────────────

def main():
    app = Application.builder().token(TELEGRAM_TOKENN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            RULE1: [CallbackQueryHandler(rule1_to_rule2,        pattern="^to_rule2$")],
            RULE2: [CallbackQueryHandler(rule2_to_rule3,        pattern="^to_rule3$")],
            RULE3: [CallbackQueryHandler(rule3_to_rule4,        pattern="^to_rule4$")],
            RULE4: [CallbackQueryHandler(rule4_to_rule5,        pattern="^to_rule5$")],
            RULE5: [CallbackQueryHandler(rule5_to_registration, pattern="^to_registration$")],
            NAME:        [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_name)],
            PHONE:       [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_phone)],
            CITIZENSHIP: [CallbackQueryHandler(handle_citizenship, pattern="^country_")],
            IIN:         [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_iin)],
            CONSENT:     [CallbackQueryHandler(handle_consent,     pattern="^consent_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )

    app.add_handler(conv_handler)
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, chat))

    print("Бот запущен...")
    app.run_polling()


if __name__ == "__main__":
    main()
